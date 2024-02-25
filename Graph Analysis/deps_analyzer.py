import os
import matplotlib.pyplot as plt
import openpyxl



class Node:
    def __init__(self, data):
        self.data = data
        self.out_neighbors = []
        self.in_neighbors = []

class Graph:

    def __init__(self, nodes, total_nodes, neighborhood_matrix):
        self.nodes = nodes
        self.total_nodes = total_nodes
        self.neighborhood_matrix = neighborhood_matrix


def parse_graph_from_file(filename):
    nodes = {}
    total_nodes = 0
    with open(filename, 'r') as file:
        for line in file:
            dest, src = map(int, line.split())
            if src not in nodes:
                nodes[src] = Node(src)
                total_nodes += 1
            if dest not in nodes:
                nodes[dest] = Node(dest)
                total_nodes += 1
            nodes[src].out_neighbors.append(nodes[dest])
            nodes[dest].in_neighbors.append(nodes[src])

    # Create the neighborhood matrix
    neighborhood_matrix = [[0] * total_nodes for _ in range(total_nodes)]
    for node in nodes.values():
        for neighbor in node.out_neighbors:
            neighborhood_matrix[node.data - 1][neighbor.data - 1] = 1

    new_graph = Graph(nodes, total_nodes, neighborhood_matrix)

    return new_graph


def calculate_degrees(nodes):
    in_degrees = {}
    out_degrees = {}
    for node in nodes.values():
        out_degrees[node.data] = len(node.out_neighbors)
        in_degrees[node.data] = len(node.in_neighbors)
    return in_degrees, out_degrees


def find_critical_path(nodes):
    sorted_nodes = topological_sort(nodes)
    distances = {node.data: float('-inf') for node in nodes.values()}
    for node in sorted_nodes:
        distances[node.data] = 0
    for node in sorted_nodes:
        for neighbor in node.out_neighbors:
            distances[neighbor.data] = max(distances[neighbor.data], distances[node.data] + 1)

    end_nodes = [node for node in nodes.values() if not node.out_neighbors]
    max_distance_node = max(end_nodes, key=lambda node: distances[node.data])
    max_distance = distances[max_distance_node.data]

    critical_path = []
    current_node = max_distance_node
    while current_node:
        critical_path.append(current_node.data)
        for neighbor in current_node.in_neighbors:
            if distances[neighbor.data] + 1 == distances[current_node.data]:
                current_node = neighbor
                break
        else:
            current_node = None

    return critical_path[::-1], max_distance


def topological_sort(nodes):
    in_degree = {node: 0 for node in nodes.values()}
    for node in nodes.values():
        for neighbor in node.out_neighbors:
            in_degree[neighbor] += 1
    queue = [node for node, degree in in_degree.items() if degree == 0]
    sorted_nodes = []

    while queue:
        current_node = queue.pop(0)
        sorted_nodes.append(current_node)
        for neighbor in current_node.out_neighbors:
            in_degree[neighbor] -= 1
            if in_degree[neighbor] == 0:
                queue.append(neighbor)

    return sorted_nodes


def count_degree_occurrences(nodes):
    in_degree_counts = {}
    out_degree_counts = {}

    for node in nodes.values():
        in_degree = len(node.in_neighbors)
        out_degree = len(node.out_neighbors)

        if in_degree in in_degree_counts:
            in_degree_counts[in_degree] += 1
        else:
            in_degree_counts[in_degree] = 1

        if out_degree in out_degree_counts:
            out_degree_counts[out_degree] += 1
        else:
            out_degree_counts[out_degree] = 1

    return in_degree_counts, out_degree_counts


def combine_degree_counts(nodes_list):
    combined_in_degree_counts = {}
    combined_out_degree_counts = {}

    for nodes in nodes_list:
        in_degree_counts, out_degree_counts = count_degree_occurrences(nodes)
        for degree, count in in_degree_counts.items():
            combined_in_degree_counts[degree] = combined_in_degree_counts.get(degree, 0) + count
        for degree, count in out_degree_counts.items():
            combined_out_degree_counts[degree] = combined_out_degree_counts.get(degree, 0) + count

    return combined_in_degree_counts, combined_out_degree_counts


def plot_degree_distribution(degree_counts, degree_type, expected_value):
    # Sort the degree counts dictionary by degree in descending order
    sorted_degrees = sorted(degree_counts.items(), key=lambda x: x[0], reverse=True)

    # Calculate the total number of nodes
    total_nodes = sum(degree_counts.values())

    # Calculate the threshold for ignoring the top 5% of degrees
    threshold = total_nodes * 0.005
    count_top_degrees = 0
    max_degree = 0
    threshold_degree = None
    # Iterate through the sorted degrees to find the maximum degree to be plotted
    for degree, count in sorted_degrees:
        count_top_degrees += count
        if count_top_degrees >= threshold:
            threshold_degree = degree
            break
        if degree > max_degree:
            max_degree = degree
    print(f"Max {degree_type} degree: {max_degree}")
    # Filter out degrees greater than the max_degree
    filtered_degrees = {degree: count for degree, count in degree_counts.items() if degree <= threshold_degree}

    # Update x and y for plotting
    x = list(filtered_degrees.keys())
    y = list(filtered_degrees.values())

    plt.bar(x, y, color='skyblue')
    plt.title(f'{degree_type} Degree Distribution (Expected Value: {expected_value})')
    plt.xlabel('Degree')
    plt.ylabel('Number of Nodes')
    plt.grid(True)
    plt.show()


def process_files_in_directory(directory):
    total_nodes_list = []
    critical_path_lengths = []
    nodes_list = []
    neighborhood_matrices = []  # Added list to store neighborhood matrices

    for filename in os.listdir(directory):
        if filename.endswith(".deps"):
            filepath = os.path.join(directory, filename)
            new_graph = parse_graph_from_file(filepath)
            nodes = new_graph.nodes
            total_nodes = new_graph.total_nodes
            neighborhood_matrix = new_graph.neighborhood_matrix
            total_nodes_list.append(total_nodes)
            nodes_list.append(nodes)
            neighborhood_matrices.append(neighborhood_matrix)

            _, length = find_critical_path(nodes)
            critical_path_lengths.append(length)

    combined_in_degree_counts, combined_out_degree_counts = combine_degree_counts(nodes_list)

    # Calculate expected value of nodes for in-degree and out-degree
    sum_in_degrees = sum(combined_in_degree_counts.values())
    if sum_in_degrees != 0:
        expected_value_in_degree_nodes = sum(
            degree * count for degree, count in combined_in_degree_counts.items()) / sum_in_degrees
    else:
        expected_value_in_degree_nodes = 0

    sum_out_degrees = sum(combined_out_degree_counts.values())
    if sum_out_degrees != 0:
        expected_value_out_degree_nodes = sum(
            degree * count for degree, count in combined_out_degree_counts.items()) / sum_out_degrees
    else:
        expected_value_out_degree_nodes = 0

    # Plotting combined in-degree distribution
    plot_degree_distribution(combined_in_degree_counts, "In", expected_value_in_degree_nodes)

    # Plotting combined out-degree distribution
    plot_degree_distribution(combined_out_degree_counts, "Out", expected_value_out_degree_nodes)

    # Calculate expected value of nodes
    expected_value_nodes = sum(total_nodes_list) / len(total_nodes_list)
    print("Expected value of nodes:", expected_value_nodes)

    # Calculate expected value of critical path length
    expected_value_critical_path_length = sum(critical_path_lengths) / len(critical_path_lengths)
    print("Expected value of critical path length:", expected_value_critical_path_length)

    return nodes_list, total_nodes_list, critical_path_lengths, neighborhood_matrices


def excel_files_in_directory(directory):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Summary"
    ws.append(["File", "Total Nodes", "Critical Path Length", "Expected In Degree", "Expected Out Degree",
               "Std Dev In Degree", "Std Dev Out Degree"])

    data_rows = []

    for filename in os.listdir(directory):
        if filename.endswith(".deps"):
            filepath = os.path.join(directory, filename)
            new_graph = parse_graph_from_file(filepath)
            nodes = new_graph.nodes
            total_nodes = new_graph.total_nodes

            in_degrees, out_degrees = calculate_degrees(nodes)
            in_degree_std_dev = calculate_standard_deviation(in_degrees)
            out_degree_std_dev = calculate_standard_deviation(out_degrees)

            expected_value_in_degree = sum(in_degrees.values()) / len(in_degrees)
            expected_value_out_degree = sum(out_degrees.values()) / len(out_degrees)

            critical_path, critical_path_length = find_critical_path(nodes)

            data_rows.append(
                [filename, total_nodes, critical_path_length, expected_value_in_degree, expected_value_out_degree,
                 in_degree_std_dev, out_degree_std_dev])

    for row in data_rows:
        ws.append(row)

    # Add Excel formula to calculate averages for each column
    last_row_index = ws.max_row
    for col in range(2, ws.max_column + 1):
        average_formula = f'=AVERAGE({openpyxl.utils.get_column_letter(col)}2:{openpyxl.utils.get_column_letter(col)}{last_row_index})'
        ws.cell(row=last_row_index + 1, column=col, value=average_formula)

    excel_filename = "data_summary.xlsx"
    wb.save(excel_filename)
    print(f"Excel file '{excel_filename}' has been created.")


def calculate_standard_deviation(degrees):
    mean = sum(degree * count for degree, count in degrees.items()) / sum(degrees.values())
    variance = sum((degree - mean) ** 2 * count for degree, count in degrees.items()) / sum(degrees.values())
    return variance ** 0.5

def transitive_closure(neighborhood_matrix):
    num_nodes = len(neighborhood_matrix)
    
    # Initialize the transitive closure matrix with the original neighborhood matrix
    closure_matrix = [row[:] for row in neighborhood_matrix]

    for k in range(num_nodes):
        for i in range(num_nodes):
            for j in range(num_nodes):
                closure_matrix[i][j] = closure_matrix[i][j] or (closure_matrix[i][k] and closure_matrix[k][j])

    return closure_matrix


deps_directory = 'deps'
process_files_in_directory(deps_directory)
excel_files_in_directory(deps_directory)
